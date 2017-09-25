# -*- coding: utf-8 -*-
"""
**************************************************************************
*                  Cross_A_Crater (e-Yantra 2016)
*                  ================================
*  This software is intended to teach image processing concepts
*	
*  UBIT Name: jayantso
*  Person Number: 50246821
*  MODULE: Project-1
*  Filename: main.py
*  Date: Sept 24, 2017
*  
*  Author: Jayant Solanki, CSE-574 Project-1, Department of Computer Science
*  and Engineering, University at Buffalo.
*  
*  Software released under Creative Commons CC BY-NC-SA
*
*  For legal information refer to:
*        http://creativecommons.org/licenses/by-nc-sa/4.0/legalcode 
*     
*
*  This software is made available on an “AS IS WHERE IS BASIS”. 
*  Licensee/end user indemnifies and will keep e-Yantra indemnified from
*  any and all claim(s) that emanate from the use of the Software or 
*  breach of the terms of this agreement.
*  
*
**************************************************************************
"""
#***********************Necessary libraries***************************#
import openpyxl as px
import numpy as np
import matplotlib.pyplot as plt
print("Ubit Name = jayantso")
print("personNumber = 50246821")

#***********************Reading from the DataSet spreadsheetll***************************#

# Loading the data from the DataSet
np.set_printoptions(precision=3)
W = px.load_workbook('DataSet/university data.xlsx', use_iterators = True)
sheet=W.get_sheet_by_name('university_data') #selecting the sheet
#getting data from the sheet and storing it into numpy array
sheetLength=sheet.max_row-1
data=np.zeros((sheetLength-1,4))#creating a zeros filled 49x4 array using numpy for storing the data from the sheet, the 4 vaiables
# print (data.shape)
# print(sheet.max_row)
# print("Reading the data from the sheet and storing into the data array...")
label=[sheet['C' + str(1)].value,sheet['D' + str(1)].value,sheet['E' + str(1)].value,sheet['F' + str(1)].value]#getting labels for the 4 columns
# print (label)
for row in range(2, sheetLength+1):
	data[row-2,0] =  (sheet['C' + str(row)].value)
	data[row-2,1] =  (sheet['D' + str(row)].value)
	data[row-2,2] =  (sheet['E' + str(row)].value)
	data[row-2,3] =  (sheet['F' + str(row)].value)

#*********************** Task 1 ***************************#
#means
mu1=np.mean(data[:,0]);
mu2=np.mean(data[:,1]);
mu3=np.mean(data[:,2]);
mu4=np.mean(data[:,3]);
#printing the means
print("mu1 = ",round(mu1,3))
print("mu2 = ",round(mu2,3))
print("mu3 = ",round(mu3,3))
print("mu4 = ",round(mu4,3))
#variance
var1=np.var(data[:,0]);
var2=np.var(data[:,1]);
var3=np.var(data[:,2]);
var4=np.var(data[:,3]);
#printing the variance
print("var1 = ",round(var1,3))
print("var2 = ",round(var2,3))
print("var3 = ",round(var3,3))
print("var4 = ",round(var4,3))
#standard deviation
sigma1=np.std(data[:,0]);
sigma2=np.std(data[:,1]);
sigma3=np.std(data[:,2]);
sigma4=np.std(data[:,3]);
#printing the sigma
print("sigma1 = ",round(sigma1,3))
print("sigma2 = ",round(sigma2,3))
print("sigma3 = ",round(sigma3,3))
print("sigma4 = ",round(sigma4,3))

#*********************** End of Task 1 ***************************#

#*********************** Task 2 ***************************#
#covariance
covarianceMat12=np.cov(data[:,0],data[:,1])
covarianceMat13=np.cov(data[:,0],data[:,2])
covarianceMat14=np.cov(data[:,0],data[:,3])
covarianceMat23=np.cov(data[:,1],data[:,2])
covarianceMat24=np.cov(data[:,1],data[:,3])
covarianceMat34=np.cov(data[:,2],data[:,3])
# covarianceMat=(data[:,0]-mu1)*(data[:,1]-mu2)/(48)
covarianceMat=np.cov(np.transpose(data))
print("covarianceMat = \n",covarianceMat)
correlationMat=np.corrcoef(np.transpose(data))
print("correlationMat = \n",correlationMat)
# plotting pairwise columns
fig1 = plt.figure()
p12 = fig1.add_subplot(311)
p12.plot(data[:,0],data[:,1], 'ro')
plt.title('CS Score (USNews) vs rest three variables')
plt.ylabel('Research Overhead %')
plt.xlabel('CS Score (USNews)')
# for xy in zip(data[:,0], data[:,1]):                                       # <--
#     p12.annotate('(%s,%s)' % xy, xy=xy, textcoords='data')

p13 = fig1.add_subplot(312)
p13.plot(data[:,0],data[:,2], 'go')
plt.ylabel('Admin Base Pay$')
plt.xlabel('CS Score (USNews)')

p14 = fig1.add_subplot(313)
p14.plot(data[:,0],data[:,3], 'bo')
plt.ylabel('Tuition(out-state)$')
plt.xlabel('CS Score (USNews)')

plt.tight_layout()
plt.show()
######################
fig2 = plt.figure()
p23 = fig2.add_subplot(211)
p23.plot(data[:,1],data[:,2], 'ro')
plt.title('Research Overhead % vs rest two variables')
plt.ylabel('Admin Base Pay$')
plt.xlabel('Research Overhead %')
# for xy in zip(data[:,0], data[:,1]):                                       # <--
#     p12.annotate('(%s,%s)' % xy, xy=xy, textcoords='data')

p24 = fig2.add_subplot(212)
p24.plot(data[:,1],data[:,3], 'go')
plt.ylabel('Tuition(out-state)$')
plt.xlabel('Research Overhead %)')


plt.tight_layout()
plt.show()
#############
fig3 = plt.figure()
p34 = fig3.add_subplot(111)
p34.plot(data[:,2],data[:,3], 'ro')
plt.title('Admin Base Pay$  vs Tuition(out-state)$')
plt.ylabel('Tuition(out-state)$')
plt.xlabel('Admin Base Pay$')
# for xy in zip(data[:,0], data[:,1]):                                       # <--
#     p12.annotate('(%s,%s)' % xy, xy=xy, textcoords='data')

plt.tight_layout()
plt.show()

